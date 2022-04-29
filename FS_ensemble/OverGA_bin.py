# official ga firstly and then oversampling ensemble
# split to five cuts before ga
# which is the formal 5 fold validations
# 先 oversampling 在 Ga
from sklearn.model_selection import StratifiedKFold
from typing_extensions import final
from Py_FS.wrapper.nature_inspired import GA
from sklearn import datasets
import numpy as np
import time
import matplotlib.pyplot as plt
import requests
from sklearn.model_selection import train_test_split
from sklearn import datasets
from sklearn.preprocessing import MinMaxScaler
from sklearn.utils.validation import column_or_1d
from scipy.spatial import distance
# from Py_FS.wrapper.nature_inspired._utilities import Solution, Data, initialize, sort_agents, display, compute_fitness, Conv_plot

# from Py_FS.wrapper.nature_inspired._utilities import Solution, Data, initialize, sort_agents, display, compute_fitness, Conv_plot
# import ga

# data = datasets.load_iris()
# d = GA(20, 100, data.data, data.target)
'''
Py FS.wrapper.nature inspired.GA(num agents, max iter, train data,
train label, obj function=compute fitness, trans function shape=‘s’, prob cross=0.4,
prob mut=0.3, save conv graph=False)
'''
# First: oversampling  Second: feature selection
from collections import Counter
import numpy as np
import smote_variants as sv
import os
import pandas as pd
import numpy as np
from sklearn.tree import DecisionTreeClassifier
import statistics
from sklearn.metrics import roc_auc_score
from sklearn import preprocessing
from sklearn.preprocessing import LabelEncoder
from imblearn.over_sampling import SMOTE
from collections import Counter
import math
from sklearn import metrics
from sklearn.cluster import KMeans
from sklearn.cluster import AffinityPropagation
from collections import Counter
from yellowbrick.cluster import KElbowVisualizer
import matplotlib.pyplot as pl
import random

from sklearn import svm, ensemble
# package
from numpy.core.fromnumeric import mean, size
from openpyxl import load_workbook
from openpyxl.styles import Font
from sklearn import tree

from itertools import permutations
import os
import numpy as np
import pandas as pd
import sys


from sklearn.tree import DecisionTreeClassifier
import statistics
from sklearn.metrics import roc_auc_score
from sklearn import svm, ensemble
import time
import datetime
import math
from sklearn.tree import DecisionTreeClassifier
import statistics
from sklearn.metrics import roc_auc_score

from imblearn.over_sampling import SMOTE
from collections import Counter
import math
from sklearn import metrics
from sklearn.cluster import KMeans
from sklearn.cluster import AffinityPropagation
from collections import Counter
from yellowbrick.cluster import KElbowVisualizer
import matplotlib.pyplot as pl
import random

from sklearn import svm, ensemble
from imblearn.over_sampling._smote.base import SMOTE
from imblearn.over_sampling import SMOTE
import smote_variants as sv
import pandas as pd
import numpy as np
from sklearn import preprocessing
from sklearn.preprocessing import LabelEncoder
import random
import math
import os
from sklearn.cluster import KMeans
from yellowbrick.cluster import KElbowVisualizer
from collections import Counter

from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import chi2

from sklearn.model_selection import cross_val_score

# for bank dataset
# 都做好前處理了 在進行 cross validation
# 前處理有以下部分：
# feature selection
# oversampling


# 前處理


def find_maj(sample_class):  # 給 class 資料
    # sample_class = labelencoder.fit_transform()
    counter = Counter(sample_class)
    maj = list(dict(counter.most_common(1)).keys())

    print('mm', type(maj))
    # maj = "".join(maj)
    maj = " ".join('%s' % id for id in maj)
    # maj = str(maj)
    print(maj)
    return maj


def classprocess(output):
    c = Counter(output)
    datagap = []
    maj = find_maj(output)
    maj = maj.replace('[]', '')
    print('flat', maj)
    print('dd', dict(c)[0])
    # maj_num = dict(c)[find_maj(output)]
    maj = int(maj)
    maj_num = dict(c)[maj]
    for className, number in c.items():
        #print(className," ",number)
     #   print(number)
        temp = np.array([className, (maj_num - number)])
        datagap.append(temp)
    return datagap


def preprocess(data):
    '''
    finaldata represent the X in the data (input atrribute)
    output repesent the y in the data (output attribute)
    '''
    le = preprocessing.LabelEncoder()
    lastColumn = data.columns[-1]
    # data[lastColumn] = data[lastColumn].str.replace(
    #     "\n", "").str.strip()
    l = data.shape[1]-1
    output = data.iloc[:, l]
    output = le.fit_transform(output)
    classCount = classprocess(output)
    finaldata = data.iloc[:, :l]
    finaldata.iloc[:, 0] = le.fit_transform(
        finaldata.iloc[:, 0])
    output = le.fit_transform(output)

    return classCount, finaldata, output


def GAselect(df):
    # df 會是資料 dataframe 模式的
    labelencoder = LabelEncoder()
    array = df.values
    X = array[:, :-1]
    Y = array[:, -1]
    Y = labelencoder.fit_transform(Y)
    X = MinMaxScaler().fit(X).transform(X)
    # Y = array[:,128]
    columnselect = GA(df, 20, 100, X, Y)
    lastcol = df.columns[-1]
    columnselect = columnselect.append(lastcol)
    dfafterselect = pd.DataFrame(df, columns=columnselect)
    return dfafterselect


def synth(finaldata, output, method):
    finaldata = np.array(finaldata)
    output = np.array(output)
    if method is 'poly':  # "poly" in method:
        print("pol")
        over = sv.polynom_fit_SMOTE()
    elif method is 'prow':  # "proW" in method:
        print("pro")
        over = sv.ProWSyn()
    elif method is 'SMOTEIPF':  # "SMOTEIPF" in method:
        print("smoteipf")
        over = sv.SMOTE_IPF()
    elif method is 'smote':
        print("smote")
        over = SMOTE(k_neighbors=2)
        X_synth, y_syth = over.fit_resample(finaldata, output)
        return X_synth, y_syth
    elif method is 'baseline':
        return finaldata, output

    X_synth, y_syth = over.sample(finaldata, output)
    return X_synth, y_syth


def ElbowCenterGenerate(train, ratio, method, path):
    # 經過 select column 完的
    # train 直接就是資料
    alloverpolynom = []
    overpolynom = []
    centerpolynom = []
    centerpolynomvalue = []
    countfor = 0

    le = preprocessing.LabelEncoder()
    classCount, finaldata, output = preprocess(train)

    minclass = Counter(output).most_common()[1][0]
    print('minclass', minclass)
    print('traincc', Counter(train.iloc[:, -1]))
    X_polynom, y_polynom = synth(
        finaldata, output, method)
    X_polynom = pd.DataFrame(X_polynom, columns=finaldata.columns)
    y_polynom = pd.DataFrame(y_polynom, columns=[train.columns[-1]])
    alloverpolynom = pd.concat(
        [X_polynom, y_polynom], axis=1)  # SMOTE 完後的數據 dataframe
    # overpolynom.append(alloverpolynom)
    print('cc', Counter(alloverpolynom.iloc[:, -1]))
    print('allover', alloverpolynom.shape)
    overpolynom = alloverpolynom  # dataframe
    tempcenterpolynom = []
    originlen = train.shape[0]
    for i in range(len(classCount)):  # 不同類個別要產生多少數據才能平衡 目前是二分類
        origincount = int(classCount[i][1])
        print("要產生的資料數", origincount)
        countfor = math.floor(
            int(classCount[i][1])*ratio)  # 要產生多少數據  無條件捨去
    #randomIndex.extend([random.randint(len(data),len(X_smote)-1) for _ in range(count)])

        if(countfor > 0):
            print(len(train.columns))
            dtemp = overpolynom
            # dtemp = pd.DataFrame(overpolynom) # overpolynom 已經是 dataframe 了
            print('dsd', dtemp.shape)
            ln = dtemp.columns[-1]
            Xtest = dtemp[dtemp[ln] == minclass]
            print('minccc', Counter(Xtest.iloc[:, -1]))
            # X = dtemp[dtemp[ln] == minclass].iloc[:,
            #                                       :dtemp.shape[1]-1]  # 選取小類的資料

            X = dtemp.iloc[originlen:, :dtemp.shape[1]-1]  # 後來生成的 都是小類
            X.reset_index(inplace=True, drop=True)
            Xcluster = np.array(X)
        # print("要產生多少",countfor)
        # 計算應該分成幾群
            model = KMeans()
            visualizer = KElbowVisualizer(
                model, k=(1, 20))  # origin
            # visualizer = kelbow_visualizer(model,Xcluster, k=(1, 12))
            # Fit the data to the visualizer
            kmodel = visualizer.fit(Xcluster)
            cluster_count = kmodel.elbow_value_  # 最佳要分成幾群
            kmeans = KMeans(n_clusters=cluster_count)
            kmeans.fit(Xcluster)
            label = Counter(kmeans.labels_)  # 標籤分類狀況

            # 不同群的比例
            labelRatio = []
            for key, element in sorted(label.items()):
                labelRatio.append(element/origincount)

            # print(X.columns[-1])
        # 把分類標籤跟原始資料進行合併
            klabel = pd.DataFrame(
                {'label': kmeans.labels_})  # 建立一個欄位名為 label 的
            df = pd.concat([X, klabel], axis=1)  # X 是後來生成的數據 類別都是小類

            print(df.columns)
            centers = kmeans.cluster_centers_  # 各群群中心
            # print(centers[0],'\nthe other',centers[1])
            # distance = []#deleted
            X = X.astype('float64')
            centers = centers.astype('float64')
            tempindata = {}
            distancesortemp = []

        # 計算每個點跟各群中心的距離

            ct = 0
        # print("分成",cluster_count,"群")
        # print("要產生",countfor)
            tempcenterpolynom = []  # 清空
            for ic in range(cluster_count):  # 總共分成幾群 每一群有其群中心
                ct += 1
                tempindata = {}  # 清空

                temppolynom = []
            # 把不同群過濾出來
                # df 是 X 跟 label 結合後的 dataframe
                # tempdf 是指那一群的資料
                tempdf = df[df['label'] == ic]
                tempdf = tempdf.iloc[:, :-1]

                print(ic, '類共有幾個data', tempdf.shape[0], 'len', len(tempdf))
            # allCluster.append(df[df['label']==ic])
                '''
                因為計算速度慢，所以嘗試使用numpy去計算
                '''
                tempdf = np.array(tempdf.values)
            # 計算每個點跟群中心的距離 速度慢
                '''
                for i in range(tempdf.shape[0]):  # 列 也就是幾筆資料

                    distance = []
                    temp = 0  # 放算出來的距離
                    tempsum = 0

                    # 到前一欄 因為最後一欄為 label
                    for j in range(tempdf.shape[1]-2):
                        # 該欄位跟center欄位的距離
                        temp = pow((centers[ic][j]-tempdf.iloc[i][j]), 2)
                        tempsum = tempsum + temp
                    # print(tempsum)
                        tempindata[i] = tempsum
                '''

                '''
                改成用 numpy 計算
                '''
                centeric = np.array(centers[ic])
                print(centeric.shape)
                print(tempdf.shape)
                for i in range(tempdf.shape[0]):  # 列 也就是幾筆資料

                    tempindata[i] = distance.euclidean(centeric, tempdf[i])

                # print('第',ic,'類',len(tempindata))
                # print('第',ic,'類',tempindata)
                # 排序跟center計算出來的距離 tempindata 是一個dict key 存的是原始的index, value存的是距離
                distancesortemp = sorted(
                    tempindata.items(), key=lambda item: item[1])
                print('算完 sort')
                # print('第',ic,'類',distancesortemp)
            # 要按照比例挑出資料

                countforlabel = math.ceil(
                    countfor * labelRatio[ic])  # 按照比例 給不同的數量 不同群不同數量
            # print("比例",labelRatio)
                temppolynom.extend(
                    distancesortemp[:countforlabel])  # 該群所要的數量
            # print("該群所要的數量",len(temppolynom))
        # tempcenterpolynom.extend(temppolynom) # 該份資料集所要的所有資料
            # print("ct",ct)
                # print('weig',temppolynom)
                tempcenterpolynom = tempcenterpolynom+temppolynom
            # print('weig22',tempcenterpolynom)
            centerpolynom.append(tempcenterpolynom)  # 所有群所要的資料集 # 所有資料集所選到的資料
            # print('centerpoly',centerpolynom)
        # print("真的有幾筆",len(centerpolynom[ii]))
    # print(centerpolynom[0])
    # print('alldatalist',list(overpolynom[0].iloc[12]))
    # print('alldata',overpolynom)
    # for i in range(len(centerpolynom)):
    #     print(i)
    alltemp = []
    for j in range(len(centerpolynom[0])):
        indexpolynom = centerpolynom[0][j][0] + originlen

        alltemp.append(list(overpolynom.iloc[indexpolynom]))
        # print('index choose',alltemp)
        # 進行 feature selection filter chi-square
    centerpolynomvalue.append(alltemp)
    print(np.shape(centerpolynomvalue))
    # 輸出的是
    return centerpolynomvalue


'''

def ElbowCenterGenerate(train, ratio, method, path):
    # 經過 select column 完的
    # train 直接就是資料
    alloverpolynom = []
    overpolynom = []
    centerpolynom = []
    centerpolynomvalue = []
    countfor = 0

    le = preprocessing.LabelEncoder()
    classCount, finaldata, output = preprocess(train)
    labelencoder = LabelEncoder()
    output = labelencoder.fit_transform(output)
    minclass = Counter(output).most_common()[1][0]
    print('minclass', minclass)
    print('traincc', Counter(train.iloc[:, -1]))
    X_polynom, y_polynom = synth(
        finaldata, output, method)
    X_polynom = pd.DataFrame(X_polynom, columns=finaldata.columns)
    y_polynom = pd.DataFrame(y_polynom, columns=[train.columns[-1]])
    alloverpolynom = pd.concat(
        [X_polynom, y_polynom], axis=1)  # SMOTE 完後的數據
    # overpolynom.append(alloverpolynom)
    print('cc', Counter(alloverpolynom.iloc[:, -1]))
    print('allover', alloverpolynom.shape)
    overpolynom = alloverpolynom
    tempcenterpolynom = []
    originlen = train.shape[0]
    for i in range(len(classCount)):  # 不同類個別要產生多少數據才能平衡 目前是二分類
        origincount = int(classCount[i][1])
        print("要產生的資料數", origincount)
        countfor = math.floor(
            int(classCount[i][1])*ratio)  # 要產生多少數據  無條件捨去
    #randomIndex.extend([random.randint(len(data),len(X_smote)-1) for _ in range(count)])

        if(countfor > 0):
            print(len(train.columns))

            dtemp = pd.DataFrame(overpolynom)
            print('dsd', dtemp.shape)
            ln = dtemp.columns[-1]
            Xtest = dtemp[dtemp[ln] == minclass]
            print('minccc', Counter(Xtest.iloc[:, -1]))
            # X = dtemp[dtemp[ln] == minclass].iloc[:,
            #                                       :dtemp.shape[1]-1]  # 選取小類的資料

            X = dtemp.iloc[originlen:, :dtemp.shape[1]-1]  # 後來生成的 都是小類
            X.reset_index(inplace=True, drop=True)
        # print("要產生多少",countfor)
        # 計算應該分成幾群
            model = KMeans()
            visualizer = KElbowVisualizer(model, k=(1, 12))

            # Fit the data to the visualizer
            kmodel = visualizer.fit(X)
            cluster_count = kmodel.elbow_value_  # 最佳要分成幾群
            kmeans = KMeans(n_clusters=cluster_count)
            kmeans.fit(X)
            label = Counter(kmeans.labels_)  # 標籤分類狀況

            # 不同群的比例
            labelRatio = []
            for key, element in sorted(label.items()):
                labelRatio.append(element/origincount)

            # print(X.columns[-1])
        # 把分類標籤跟原始資料進行合併
            klabel = pd.DataFrame(
                {'label': kmeans.labels_})  # 建立一個欄位名為 label 的
            df = pd.concat([X, klabel], axis=1)  # X 是後來生成的數據 類別都是小類

            print(df.columns)
            centers = kmeans.cluster_centers_  # 各群群中心
            # print(centers[0],'\nthe other',centers[1])
            distance = []
            X = X.astype('float64')
            centers = centers.astype('float64')
            tempindata = {}
            distancesortemp = []

        # 計算每個點跟各群中心的距離

            ct = 0
        # print("分成",cluster_count,"群")
        # print("要產生",countfor)
            tempcenterpolynom = []  # 清空
            for ic in range(cluster_count):  # 總共分成幾群 每一群有其群中心
                ct += 1
                tempindata = {}  # 清空

                temppolynom = []
            # 把不同群過濾出來
                # df 是 X 跟 label 結合後的 dataframe
                # tempdf 是指那一群的資料
                tempdf = df[df['label'] == ic]

                print(ic, '類共有幾個data', tempdf.shape[0], 'len', len(tempdf))
            # allCluster.append(df[df['label']==ic])

            # 計算每個點跟群中心的距離
                for i in range(tempdf.shape[0]):  # 列 也就是幾筆資料

                    distance = []
                    temp = 0  # 放算出來的距離
                    tempsum = 0
                    # 到前一欄 因為最後一欄為 label
                    for j in range(tempdf.shape[1]-2):
                        # 該欄位跟center欄位的距離
                        temp = pow((centers[ic][j]-tempdf.iloc[i][j]), 2)
                        tempsum = tempsum + temp
                    # print(tempsum)
                        tempindata[i] = tempsum
                # print('第',ic,'類',len(tempindata))
                # print('第',ic,'類',tempindata)
                # 排序跟center計算出來的距離 tempindata 是一個dict key 存的是原始的index, value存的是距離
                distancesortemp = sorted(
                    tempindata.items(), key=lambda item: item[1])
                # print('第',ic,'類',distancesortemp)
            # 要按照比例挑出資料

                countforlabel = math.ceil(
                    countfor * labelRatio[ic])  # 按照比例 給不同的數量 不同群不同數量
            # print("比例",labelRatio)
                temppolynom.extend(
                    distancesortemp[:countforlabel])  # 該群所要的數量
            # print("該群所要的數量",len(temppolynom))
        # tempcenterpolynom.extend(temppolynom) # 該份資料集所要的所有資料
            # print("ct",ct)
                # print('weig',temppolynom)
                tempcenterpolynom = tempcenterpolynom+temppolynom
            # print('weig22',tempcenterpolynom)
            centerpolynom.append(tempcenterpolynom)  # 所有群所要的資料集 # 所有資料集所選到的資料
            # print('centerpoly',centerpolynom)
        # print("真的有幾筆",len(centerpolynom[ii]))
    # print(centerpolynom[0])
    # print('alldatalist',list(overpolynom[0].iloc[12]))
    # print('alldata',overpolynom)
    # for i in range(len(centerpolynom)):
    #     print(i)
    alltemp = []
    for j in range(len(centerpolynom[0])):
        indexpolynom = centerpolynom[0][j][0] + originlen

        alltemp.append(list(overpolynom.iloc[indexpolynom]))
        # print('index choose',alltemp)
        # 進行 feature selection filter chi-square
    centerpolynomvalue.append(alltemp)
    print(np.shape(centerpolynomvalue))
    # 輸出的是
    return centerpolynomvalue
'''


def calculatethreemethod(train, approach, sheetName, path, test_X, test_y):
    '''
    train 是讀完的資料
    '''
    cell = 1

    # originpath = os.getcwd()
    # os.chdir(path)
    # tempfileName = train[0]  # 要擷取 file 的名字
    # tempfileName = tempfileName.split('-')
    # fileName = '-'.join(tempfileName[:-2])
    # data = pd.read_excel(train[0], index_col=0)

    # print("dsdsadad", data.columns)
    tempappr = []  # 放要比較的數值(3,3,4 之間會比較)
    randompoly = []
    randomPro = []
    randomIPF = []
    start = 0
    end = 0
    # traindata = pd.read_excel(train,index_col=0)

    # data = traindata.iloc[:, train.shape[1]-1]
    # target = traindata.iloc[:, train.shape[1]-1]
    # columnchoose = GA(traindata, 20, 100, data, target)
    # traindata = pd.DataFrame(traindata, columns=columnchoose)
    accuracyDe = []
    for index in range(len(f)):
        # train 直接就是資料
        randompoly = ElbowCenterGenerate(
            train, int(f[index][0])*0.1, "poly", path)

        randomPro = ElbowCenterGenerate(
            train, int(f[index][1])*0.1, "prow", path)
        randomIPF = ElbowCenterGenerate(
            train, int(f[index][2])*0.1, "SMOTEIPF", path)

        # start = time.process_time()

        allRandom = []
        allRandom = randompoly[0] + randomPro[0] + randomIPF[0]

        temp = []
        # print('elbow',randomPro.shape)
        # for i in range(len(randomPro)):
        #     temp = randomIPF[i] + randomPro[i]  # list 合併
        #     temp = temp + randompoly[i]
        #     temp = np.array(temp)
        #     allRandom.append(temp)

        # print('feature:', allcolumns)
        # ==============================
        '''先 wrapper 再 oversampling'''
        # for j in range(len(allRandom)):  #
        #     allRandom[j] = pd.DataFrame(allRandom[j], columns=columnchoose)

        print('following123\n', np.shape(allRandom))
        allRandom = pd.DataFrame(allRandom, columns=train.columns)
        cell = cell + 1
        # 為了算個別 C4.5 跟 SVM 時間 所以先註解掉 SVM 的部分
        # temptime = time.process_time()
        mergedata = pd.concat([allRandom, train], axis=0)
        # return mergedata
        dfafterselect = GAselect(mergedata)  # 加入 GA 的部分
        print(mergedata)
        # meanDe = foldval(mergedata)
        train_X, train_y = dfafterselect.iloc[:,
                                              :-1], dfafterselect.iloc[:, -1]
        clf = DecisionTreeClassifier().fit(train_X, train_y)

        # clf = svm.SVC(kernel='rbf', C=1).fit(train_X, train_y)  # linear rbf
        # clf = svm.SVC(kernel='linear', C=1, gamma='auto').fit(train_X, train_y)
        test_y_predicted = clf.predict(test_X)
        # tempmean = roc_auc_score(test_y, test_y_predicted)
        accuracyDe.append(roc_auc_score(test_y, test_y_predicted))
    print(accuracyDe)
    accstr = "  ".join([str(_) for _ in accuracyDe])
    message = '準確率 ' + accstr
    lineNotifyMessage(token, message)
    return accuracyDe
    '''
        meanDe = statistics.mean(accuracyDe)
        wb = load_workbook(
            '/Users/emily/Desktop/Research/fs_ensemble/WrapperOver.xlsx')
        sheet = wb[sheetName]
        sheet.cell(row=1, column=cell, value=meanDe)
        wb.save('/Users/emily/Desktop/Research/fs_ensemble/WrapperOver.xlsx')
    '''

    # os.chdir(originpath)


# def predictDe(train, test, allRandomHalf):

'''
def foldval(df):

    X = np.array(df.iloc[:, :-1])
    y = np.array(df.iloc[:, -1])

    kfold = StratifiedKFold(n_splits=5, shuffle=True, random_state=1)
    # enumerate the splits and summarize the distributions

    accuracyDe = []
    for train_ix, test_ix in kfold.split(X, y):
        # select rows
        train_X, test_X = X[train_ix], X[test_ix]
        train_y, test_y = y[train_ix], y[test_ix]
        # summarize train and test composition 兩個不同類別的比例
        train_0, train_1 = len(train_y[train_y == 0]), len(
            train_y[train_y == 1])
        test_0, test_1 = len(test_y[test_y == 0]), len(test_y[test_y == 1])
        print('>Train: 0=%d, 1=%d, Test: 0=%d, 1=%d' %
              (train_0, train_1, test_0, test_1))
        # 每一等分訓練都拿去做 preprocessing

        # clf = DecisionTreeClassifier().fit(train_X, train_y)

        clf = svm.SVC(kernel='rbf', C=1).fit(train_X, train_y)  # linear
        test_y_predicted = clf.predict(test_X)

        accuracyDe.append(roc_auc_score(test_y, test_y_predicted))
    print(accuracyDe)
    meanDe = statistics.mean(accuracyDe)
    return meanDe
'''
# line 推播通知


def lineNotifyMessage(token, msg):

    headers = {
        "Authorization": "Bearer " + token,
        "Content-Type": "application/x-www-form-urlencoded"
    }

    payload = {'message': msg}
    r = requests.post("https://notify-api.line.me/api/notify",
                      headers=headers, params=payload)
    return r.status_code


def main(datapath):
    token = 'bCpvlorCZoJPxA2thceajfMrdgwZhumsVTYMMOGXg4q'
    if 'xlsx' in datapath:
        df = pd.read_excel(datapath)  # 得到原始 data
    elif 'csv' in datapath:
        df = pd.read_csv(datapath, index_col=0)
    print('原始資料', df)
    # dfafterselect = GAselect(df)  # 先 feature selection 得到 feature 後的 data
    # print('經過 feature selection 後 ', dfafterselect.columns)
    # calculatethreemethod(
    #     dfafterselect, 'approach', 'firstwrappe', '')  # elbowmethod 三種

    # 切成 5 等分
    kfold = StratifiedKFold(n_splits=5, shuffle=True, random_state=1)
    # enumerate the splits and summarize the distributions
    X = np.array(df.iloc[:, :-1])
    X = MinMaxScaler().fit(X).transform(X)
    y = np.array(df.iloc[:, -1])
    accuracyDe = []
    sumaccu = np.zeros(12)
    count = 1
    for train_ix, test_ix in kfold.split(X, y):
        # select rows
        message = '開始了第' + str(count) + ' 分資料'
        count += 1
        lineNotifyMessage(token, message)
        train_X, test_X = X[train_ix], X[test_ix]
        train_y, test_y = y[train_ix], y[test_ix]
        # summarize train and test composition 兩個不同類別的比例
        train_0, train_1 = len(train_y[train_y == 0]), len(
            train_y[train_y == 1])
        test_0, test_1 = len(test_y[test_y == 0]), len(test_y[test_y == 1])
        print('>Train: 0=%d, 1=%d, Test: 0=%d, 1=%d' %
              (train_0, train_1, test_0, test_1))
        # 每一等分訓練都拿去做 preprocessing

        train_y = pd.DataFrame(train_y, columns=[df.columns[-1]])
        train_X = pd.DataFrame(train_X, columns=df.columns[:-1])
        traindf = pd.concat([train_X, train_y], axis=1)

        # dfafterselect = GAselect(traindf)
        # print('經過 feature selection 後 ', dfafterselect.columns)
        temp = np.array(calculatethreemethod(
            traindf, 'approach', 'firstwrappe', '', test_X, test_y))
        sumaccu = np.sum([sumaccu, temp], axis=0)
        accuracyDe.append([temp])
        msg = ' '.join([str(_) for _ in sumaccu])
        msg = 'sum' + str(count)+' ' + msg
        lineNotifyMessage(token, msg)
        # mergedata = calculatethreemethod(
        # dfafterselect, 'approach', 'firstwrappe', '')  # elbowmethod 三種

        # train_X = np.array(mergedata.iloc[:, :-1])
        # train_y = np.array(mergedata.iloc[:, -1])

    #     clf = DecisionTreeClassifier().fit(train_X, train_y)

    #     # clf = svm.SVC(kernel='rbf', C=1).fit(train_X, train_y)  # linear
    #     test_y_predicted = clf.predict(test_X)

    #     accuracyDe.append(roc_auc_score(test_y, test_y_predicted))
    # print(accuracyDe)
    # meanDe = statistics.mean(accuracyDe)

    sumaccu = sumaccu/5
    print(sumaccu)
    msg = ' '.join([str(_) for _ in sumaccu])
    lineNotifyMessage(token, msg)

    cell = 1
    for i in sumaccu:
        wb = load_workbook(
            '/Users/emily/Desktop/Research/fs_ensemble/WrapperOver.xlsx')
        sheet = wb['firstover']
        sheet.cell(row=1, column=cell, value=i)
        cell += 1
        wb.save('/Users/emily/Desktop/Research/fs_ensemble/WrapperOver.xlsx')
    return 0


if __name__ == '__main__':
    f1 = list(permutations("442", 3))
    f2 = list(permutations("253", 3))
    f3 = list(permutations("334", 3))

    f = f1+f2+f3
    temp = []
    # get rid of the repeat ratio composition
    for i in f:
        if i not in temp:
            temp.append(i)
    f = temp
    token = 'bCpvlorCZoJPxA2thceajfMrdgwZhumsVTYMMOGXg4q'
    lineNotifyMessage(
        token, 'c4.5 hiva_agnostic + 先oversampling ensemble + GA')
    meanDe = main(
        '/Users/emily/Desktop/hiva_agnostic.csv')
