# First: oversampling  Second: feature selection
from collections import Counter
import numpy as np
import smote_variants as sv
import os
import pandas as pd
import numpy as np
from sklearn.tree import DecisionTreeClassifier
import statistics
from sklearn.metrics import roc_auc_score
from sklearn import preprocessing
from sklearn.preprocessing import LabelEncoder
from imblearn.over_sampling import SMOTE
from collections import Counter
import math
from sklearn import metrics
from sklearn.cluster import KMeans
from sklearn.cluster import AffinityPropagation
from collections import Counter
from yellowbrick.cluster import KElbowVisualizer
import matplotlib.pyplot as pl
import random
from sklearn import preprocessing
from sklearn.preprocessing import LabelEncoder
from sklearn import svm, ensemble
# package
from numpy.core.fromnumeric import size
from openpyxl import load_workbook
from openpyxl.styles import Font
from sklearn import tree

from itertools import permutations
import os
import numpy as np
import pandas as pd
import sys
from sklearn import preprocessing
from sklearn.preprocessing import LabelEncoder

from sklearn.tree import DecisionTreeClassifier
import statistics
from sklearn.metrics import roc_auc_score
from sklearn import svm, ensemble
import time
import datetime
import math
from sklearn.tree import DecisionTreeClassifier
import statistics
from sklearn.metrics import roc_auc_score
from sklearn import preprocessing
from sklearn.preprocessing import LabelEncoder
from imblearn.over_sampling import SMOTE
from collections import Counter
import math
from sklearn import metrics
from sklearn.cluster import KMeans
from sklearn.cluster import AffinityPropagation
from collections import Counter
from yellowbrick.cluster import KElbowVisualizer
import matplotlib.pyplot as pl
import random
from sklearn import preprocessing
from sklearn.preprocessing import LabelEncoder
from sklearn import svm, ensemble
from imblearn.over_sampling._smote.base import SMOTE
from imblearn.over_sampling import SMOTE
import smote_variants as sv
import pandas as pd
import numpy as np
from sklearn import preprocessing
from sklearn.preprocessing import LabelEncoder
import random
import math
import os
from sklearn.cluster import KMeans
from yellowbrick.cluster import KElbowVisualizer
from collections import Counter
from sklearn import preprocessing
from sklearn.preprocessing import LabelEncoder
from sklearn.feature_selection import SelectKBest
from sklearn.feature_selection import chi2

# dfa['Name ']


def synth(finaldata, output, method):
    finaldata = np.array(finaldata)
    output = np.array(output)
    if method is 'poly':  # "poly" in method:
        print("pol")
        over = sv.polynom_fit_SMOTE()
    elif method is 'prow':  # "proW" in method:
        print("pro")
        over = sv.ProWSyn()
    elif method is 'SMOTEIPF':  # "SMOTEIPF" in method:
        print("smoteipf")
        over = sv.SMOTE_IPF()
    elif method is 'smote':
        print("smote")
        over = SMOTE(k_neighbors=2)
        X_synth, y_syth = over.fit_resample(finaldata, output)
        return X_synth, y_syth
    elif method is 'baseline':
        return finaldata, output

    X_synth, y_syth = over.sample(finaldata, output)
    return X_synth, y_syth


# 計算要補多少值


def find_maj(sample_class):  # 給 class 資料
    counter = Counter(sample_class)
    maj = list(dict(counter.most_common(1)).keys())
    maj = "".join(maj)
    # print(maj)
    return maj


def classprocess(output):
    c = Counter(output)
    datagap = []
    maj = find_maj(output)
    maj_num = dict(c)[find_maj(output)]
    for className, number in c.items():
        #print(className," ",number)
     #   print(number)
        temp = np.array([className, (maj_num - number)])
        datagap.append(temp)
    return datagap


def preprocess(data):
    '''
    finaldata represent the X in the data (input atrribute)
    output repesent the y in the data (output attribute)
    '''
    le = preprocessing.LabelEncoder()
    lastColumn = data.columns[-1]
    data[lastColumn] = data[lastColumn].str.replace(
        "\n", "").str.strip()
    l = data.shape[1]-1
    output = data.iloc[:, l]
    classCount = classprocess(output)
    finaldata = data.iloc[:, :l]
    finaldata.iloc[:, 0] = le.fit_transform(
        finaldata.iloc[:, 0])
    output = le.fit_transform(output)

    return classCount, finaldata, output

# # 跟原始資料合併 predict with Decision & SVM method


def predictDe(train, test, allRandomHalf):

    mergeRandom = []
    accuracies = []
    le = preprocessing.LabelEncoder()
    for index, element in enumerate(train):

        data = pd.read_excel(element, index_col=0)
        l = data.shape[1]-1

        lastColumn = data.columns[-1]
        data[lastColumn] = data[lastColumn].str.replace("\n", "").str.strip()
        data.iloc[:, -1] = le.fit_transform(data.iloc[:, -1])
        data.iloc[:, 0] = le.fit_transform(data.iloc[:, 0])

        for j in range(data.shape[1]):
            for k in range(data.shape[0]):
                # print(df.iloc[j,i])
                if isinstance(data.iloc[k, j], str):
                    data.iloc[:, j] = data.iloc[:, j].apply(
                        lambda col: str(col))
                    data.iloc[:, j] = le.fit_transform(
                        data.iloc[:, j])
                else:
                    data.iloc[:, j] = le.fit_transform(
                        data.iloc[:, j])
                    break

        df = data
        finaldata = df.iloc[:, :l]
        output = df.iloc[:, l]
        feature_count = df.shape[1]-1
        # X,y = df.iloc[:,:-1],df.iloc[:,-1]
        countforfeature = int(feature_count * 0.8)
        chimodel = SelectKBest(score_func=chi2, k=countforfeature)
        fit = chimodel.fit(finaldata, output)
        # X_new = chimodel.fit_transform(X)
        finaldata = fit.transform(finaldata)
        finaldata = pd.DataFrame(finaldata)
        output.reset_index(inplace=True, drop=True)
        print('finaldata:\n', finaldata)
        print('output\n', output)
        print('merge前 con', Counter(data.iloc[:, -1]))
        print('final ', finaldata.shape, 'output ', output.shape)
        data = pd.concat([finaldata, output], axis=1)
        print('merge後 con', Counter(data.iloc[:, -1]))
        print('data', data.columns)
        print('allrandom', allRandomHalf[index].columns)
        data.columns = allRandomHalf[index].columns
        print('mergedata', data.columns)  # 15
        print('allRadomHalf\n', allRandomHalf[index].columns)
        # print(allRandomHalf.isnull())
        # lastColumn = data.columns[-1]

        # data[lastColumn] = data[lastColumn].str.replace("\n", "").str.strip()

        mergeRandom = pd.concat([data, allRandomHalf[index]], axis=0)
        print('mr', mergeRandom.shape)
        # finaldata = mergeRandom.iloc[:, :feature_count]
        # output = mergeRandom.iloc[:, feature_count]

        # 把非 numeric 的資料用 label encoder 轉成 numeric 資料
        # for j in range(mergeRandom.shape[1]):
        #     for k in range(mergeRandom.shape[0]):
        #         # print(df.iloc[j,i])
        #         if isinstance(mergeRandom.iloc[k, j], str):
        #             mergeRandom.iloc[:, j] = mergeRandom.iloc[:, j].apply(
        #                 lambda col: str(col))
        #             mergeRandom.iloc[:, j] = le.fit_transform(
        #                 mergeRandom.iloc[:, j])
        #             break
        # print('allrandom\n',allRandomHalf.shape)

        finaldata = mergeRandom.iloc[:, :-1]
        output = mergeRandom.iloc[:, -1]
        print(Counter(output))

        clf = DecisionTreeClassifier()
        clf = clf.fit(finaldata, output)

        # 不然會有多出來的 unnamed column
        test_file = pd.read_excel(test[index], index_col=0)
        test_data = pd.DataFrame(test_file)

        test_X = test_data.iloc[:, :(test_data.shape[1])-1]
        # 把非 numeric 的資料用 label encoder 轉成 numeric 資料
        for j in range(test_X.shape[1]):
            for k in range(test_X.shape[0]):
                # print(df.iloc[j,i])
                if isinstance(test_X.iloc[k, j], str):
                    test_X.iloc[:, j] = test_X.iloc[:, j].apply(
                        lambda col: str(col))
                    test_X.iloc[:, j] = le.fit_transform(test_X.iloc[:, j])
                else:
                    test_X.iloc[:, j] = le.fit_transform(test_X.iloc[:, j])
                    break
        test_X.iloc[:, 0] = le.fit_transform(test_X.iloc[:, 0])
        test_y = test_data.iloc[:, test_data.shape[1]-1]

        fit = chimodel.fit(test_X, test_y)
        # X_new = chimodel.fit_transform(X)
        test_X = fit.transform(test_X)

        test_y_predicted = clf.predict(test_X)

        test_y = le.fit_transform(test_y)
        test_y_predicted = le.fit_transform(test_y_predicted)

        accuracy = roc_auc_score(test_y, test_y_predicted)
        accuracies.append(accuracy)

    mean = statistics.mean(accuracies)
    mean = statistics.mean(accuracies)
    meanRound = round(mean, 3)
    print(meanRound)
    return meanRound


def predictSVM(train, test, allRandomHalf):
    mergeRandom = []
    accuracies = []
    le = preprocessing.LabelEncoder()
    for index, element in enumerate(train):
        data = pd.read_excel(element, index_col=0)
        lastColumn = data.columns[-1]

        data[lastColumn] = data[lastColumn].str.replace("\n", "").str.strip()
        l = data.shape[1]-1

        data.iloc[:, l] = le.fit_transform(data.iloc[:, l])
        data.iloc[:, 0] = le.fit_transform(data.iloc[:, 0])
        for j in range(data.shape[1]):
            for k in range(data.shape[0]):
                # print(df.iloc[j,i])
                if isinstance(data.iloc[k, j], str):
                    data.iloc[:, j] = data.iloc[:, j].apply(
                        lambda col: str(col))
                    data.iloc[:, j] = le.fit_transform(
                        data.iloc[:, j])
                else:
                    data.iloc[:, j] = le.fit_transform(
                        data.iloc[:, j])
                    break

        df = data
        finaldata = df.iloc[:, :l]
        output = df.iloc[:, l]
        feature_count = df.shape[1]-1
        # X,y = df.iloc[:,:-1],df.iloc[:,-1]
        countforfeature = int(feature_count * 0.8)
        chimodel = SelectKBest(score_func=chi2, k=countforfeature)
        fit = chimodel.fit(finaldata, output)
        # X_new = chimodel.fit_transform(X)
        finaldata = fit.transform(finaldata)
        finaldata = pd.DataFrame(finaldata)
        output.reset_index(inplace=True, drop=True)
        # print('finaldata:\n', finaldata)
        # print('output\n', output)
        # print('merge前 con', Counter(data.iloc[:, -1]))
        # print('final ', finaldata.shape, 'output ', output.shape)
        data = pd.concat([finaldata, output], axis=1)

        data.columns = allRandomHalf[index].columns

        # print(allRandomHalf.isnull())
        mergeRandom = pd.concat([data, allRandomHalf[index]], axis=0)

        finaldata = mergeRandom.iloc[:, :-1]
        output = mergeRandom.iloc[:, -1]
        print(Counter(output))

        clf = svm.SVC(kernel='linear', C=1, gamma='auto')  # 要改成 linear
        clf = clf.fit(finaldata, output)

        # 不然會有多出來的 unnamed column
        test_file = pd.read_excel(test[index], index_col=0)
        test_data = pd.DataFrame(test_file)
        test_X = test_data.iloc[:, :(test_data.shape[1])-1]

        # 把非 numeric 的資料用 label encoder 轉成 numeric 資料
        for j in range(test_X.shape[1]):
            for k in range(test_X.shape[0]):
                # print(df.iloc[j,i])
                if isinstance(test_X.iloc[k, j], str):
                    test_X.iloc[:, j] = test_X.iloc[:, j].apply(
                        lambda col: str(col))
                    test_X.iloc[:, j] = le.fit_transform(test_X.iloc[:, j])
                else:
                    test_X.iloc[:, j] = le.fit_transform(test_X.iloc[:, j])
                    break

        test_X.iloc[:, 0] = le.fit_transform(test_X.iloc[:, 0])
        test_y = test_data.iloc[:, test_data.shape[1]-1]

        test_y = le.fit_transform(test_y)

        fit = chimodel.fit(test_X, test_y)
        test_X = fit.transform(test_X)

        test_y_predicted = clf.predict(test_X)

        test_y_predicted = le.fit_transform(test_y_predicted)

        accuracy = roc_auc_score(test_y, test_y_predicted)
        accuracies.append(accuracy)

    mean = statistics.mean(accuracies)
    meanRound = round(mean, 3)
    print(meanRound)
    return meanRound


##############

# HERE
def ElbowCenterGenerate(train, ratio, method, path):
    if ratio == 0:
        return []
    alloverpolynom = []
    overpolynom = []
    centerpolynom = []
    centerpolynomvalue = []
    countfor = 0
    le = preprocessing.LabelEncoder()
    for ii, i in enumerate(train):
        print("第幾個", i)
        data = pd.read_excel(i, index_col=0)
        #print(i, "traindata1", data)

        classCount, finaldata, output = preprocess(data)
        # 把非 numeric 的資料用 label encoder 轉成 numeric 資料
        for j in range(finaldata.shape[1]):
            for k in range(finaldata.shape[0]):
                # print(df.iloc[j,i])
                if isinstance(finaldata.iloc[k, j], str):
                    finaldata.iloc[:, j] = finaldata.iloc[:, j].apply(
                        lambda col: str(col))
                    finaldata.iloc[:, j] = le.fit_transform(
                        finaldata.iloc[:, j])
                else:
                    finaldata.iloc[:, j] = le.fit_transform(
                        finaldata.iloc[:, j])
                    break

        print(Counter(output))
        originlen = data.shape[0]  # 原始的 data 數量
        X_polynom, y_polynom = synth(
            finaldata, output, method)
        X_polynom = pd.DataFrame(X_polynom)
        y_polynom = pd.DataFrame(y_polynom)
        alloverpolynom = pd.concat(
            [X_polynom, y_polynom], axis=1)  # SMOTE 完後的數據
        overpolynom.append(alloverpolynom)
        tempcenterpolynom = []
        for i in range(len(classCount)):  # 不同類個別要產生多少數據才能平衡 目前是二分類
            origincount = int(classCount[i][1])
            print("要產生的資料數", origincount)
            countfor = math.floor(
                int(classCount[i][1])*ratio)  # 要產生多少數據  無條件捨去
        #randomIndex.extend([random.randint(len(data),len(X_smote)-1) for _ in range(count)])

            if(countfor > 0):
                dtemp = pd.DataFrame(overpolynom[ii])
                X = dtemp.iloc[originlen:, :dtemp.shape[1]-1]  # 後來生成的 都是小類
                X.reset_index(inplace=True, drop=True)
            # print("要產生多少",countfor)
            # 計算應該分成幾群
                model = KMeans()
                visualizer = KElbowVisualizer(model, k=(1, 12))

                # Fit the data to the visualizer
                kmodel = visualizer.fit(X)
                cluster_count = kmodel.elbow_value_  # 最佳要分成幾群
                kmeans = KMeans(n_clusters=cluster_count)
                kmeans.fit(X)
                label = Counter(kmeans.labels_)  # 標籤分類狀況

                # 不同群的比例
                labelRatio = []
                for key, element in sorted(label.items()):
                    labelRatio.append(element/origincount)
            # print(labelRatio)

            # 把分類標籤跟原始資料進行合併
                klabel = pd.DataFrame(
                    {'label': kmeans.labels_})  # 建立一個欄位名為 label 的
                df = pd.concat([X, klabel], axis=1)  # X 是後來生成的數據 類別都是小類
            # print(df)
                centers = kmeans.cluster_centers_  # 各群群中心

                distance = []
                X = X.astype('float64')
                centers = centers.astype('float64')
                tempindata = {}
                distancesortemp = []

            # 計算每個點跟各群中心的距離

                ct = 0
            # print("分成",cluster_count,"群")
            # print("要產生",countfor)
                tempcenterpolynom = []  # 清空
                for ic in range(cluster_count):
                    ct += 1

                    temppolynom = []
                # 把不同群過濾出來
                    # df 是 X 跟 label 結合後的 dataframe
                    tempdf = df[df['label'] == ic]
                # allCluster.append(df[df['label']==ic])

                # 計算每個點跟群中心的距離
                    for i in range(tempdf.shape[0]-1):  # 列 也就是幾筆資料

                        distance = []
                        temp = 0  # 放算出來的距離
                        tempsum = 0
                        # 到前一欄 因為最後一欄為 label
                        for j in range(tempdf.shape[1]-2):
                            # 該欄位跟center欄位的距離
                            temp = pow((centers[ic][j]-tempdf.iloc[i][j]), 2)
                            tempsum = tempsum + temp
                        # print(tempsum)
                            tempindata[i] = tempsum

                    distancesortemp = sorted(
                        tempindata.items(), key=lambda item: item[1])
                # print(distancesortemp)

                # 要按照比例挑出資料

                    countforlabel = math.ceil(
                        countfor * labelRatio[ic])  # 按照比例 給不同的數量 不同群不同數量
                # print("比例",labelRatio)
                    temppolynom.extend(
                        distancesortemp[:countforlabel])  # 該群所要的數量
                # print("該群所要的數量",len(temppolynom))
            # tempcenterpolynom.extend(temppolynom) # 該份資料集所要的所有資料

                # print("ct",ct)
                    tempcenterpolynom = tempcenterpolynom+temppolynom

                centerpolynom.append(tempcenterpolynom)  # 所有資料集所選到的資料
            # print("真的有幾筆",len(centerpolynom[ii]))
        # print(centerpolynom[0])
        for i in range(len(centerpolynom)):
            alltemp = []
            for j in range(len(centerpolynom[i])):
                indexpolynom = centerpolynom[i][j][0] + originlen - 1
                alltemp.append(list(overpolynom[i].iloc[indexpolynom]))
             # 進行 feature selection filter chi-square
            centerpolynomvalue.append(alltemp)

    return centerpolynomvalue  # , allcolumns


# use the feature selection elbow center function
# def calculatethreemethod(train, test, id, path, approach, maxsingle, sheetName):
def calculatethreemethod(train, test, id, path, approach, sheetName):
    '''
    1. train is assign by train data
    2. test
    3. approach means the choice of random or center
    4. id is for the purpose of writing into the excel different
        dataset need to be on different row
    5. path right now is useless
    '''
    cell = 1

    originpath = os.getcwd()
    os.chdir(path)
    tempfileName = train[0]  # 要擷取 file 的名字
    tempfileName = tempfileName.split('-')
    fileName = '-'.join(tempfileName[:-2])
    data = pd.read_excel(train[0], index_col=0)
    # print("dsdsadad", data.columns)
    tempappr = []  # 放要比較的數值(3,3,4 之間會比較)
    randompoly = []
    randomPro = []
    randomIPF = []
    start = 0
    end = 0
    for index in range(len(f)):
        '''
        if approach == 'random':
            randompoly = RandomCal.RandomGenerate(
                train, int(f[index][0])*0.1, "poly", path)
            randomPro = RandomCal.RandomGenerate(
                train, int(f[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.RandomGenerate(
                train, int(f[index][2])*0.1, "SMOTEIPF", path)
        elif approach == 'center':
            randompoly = RandomCal.CenterGenerate(
                train, int(f[index][0])*0.1, "poly", path)
            randomPro = RandomCal.CenterGenerate(
                train, int(f[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.CenterGenerate(
                train, int(f[index][2])*0.1, "SMOTEIPF", path)
        elif approach == 'elbowRandom':
            randompoly = RandomCal.ElbowRandomGenerate(
                train, int(f[index][0])*0.1, "poly", path)
            randomPro = RandomCal.ElbowRandomGenerate(
                train, int(f[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.ElbowRandomGenerate(
                train, int(f[index][2])*0.1, "SMOTEIPF", path)
        elif approach == 'elbowCenter':
            randompoly = RandomCal.ElbowCenterGenerate(
                train, int(f[index][0])*0.1, "poly", path)
            randomPro = RandomCal.ElbowCenterGenerate(
                train, int(f[index][1])*0.1, "prow", path)
            randomIPF = RandomCal.ElbowCenterGenerate(
                train, int(f[index][2])*0.1, "SMOTEIPF", path)
        '''

        randompoly = ElbowCenterGenerate(
            train, int(f[index][0])*0.1, "poly", path)

        randomPro = ElbowCenterGenerate(
            train, int(f[index][1])*0.1, "prow", path)
        randomIPF = ElbowCenterGenerate(
            train, int(f[index][2])*0.1, "SMOTEIPF", path)

        # start = datetime.datetime.now()
        start = time.process_time()

        allRandom = []
        temp = []
        # print('elbow',randomPro.shape)
        for i in range(len(randomPro)):
            temp = randomIPF[i] + randomPro[i]  # list 合併
            temp = temp + randompoly[i]
            temp = np.array(temp)
            allRandom.append(temp)

        # print('feature:', allcolumns)
        '''先 oversampling 再 feature selection'''
        for j in range(len(allRandom)):  #
            allRandom[j] = pd.DataFrame(allRandom[j], columns=data.columns)
            # 進行 feature selection filter chi-square
            df = allRandom[j]
            finaldata = allRandom[j].iloc[:, :-1]
            output = allRandom[j].iloc[:, -1]
            feature_count = df.shape[1]-1
            # X,y = df.iloc[:,:-1],df.iloc[:,-1]
            print('總共有：', feature_count)
            countforfeature = int(feature_count * 0.8)
            print('要篩出：', countforfeature)

            X = finaldata
            chimodel = SelectKBest(score_func=chi2, k=countforfeature)  # )
            # fitmodel = chimodel.fit(finaldata, output)
            # X_new = chimodel.fit_transform(X)
            finaldata = chimodel.fit_transform(finaldata, output)
            finaldata = pd.DataFrame(finaldata)

            # 找出挑到的 feature
            scores = chimodel.scores_
            # print('model scores:', scores) #得分越高,特征越重要
            p_values = chimodel.pvalues_
            # print('model p-values', p_values)#p-values越小,置信度越高,特征越重要
            # 按重要性排序,选出最重要的k个
            indices = np.argsort(scores)[::-1]
            # print('in',indices)
            k_best_features = list(
                X.columns.values[indices[0:countforfeature]])
            # print('k best features are: ',k_best_features)
            allcolumns = k_best_features
            allcolumns.append(data.columns[-1])
            allRandom[j] = allRandom[j][allcolumns]
            # print('allcol:',allcolumns)

        print('following123\n', allRandom[0].shape)

        cell = cell + 1
        # 為了算個別 C4.5 跟 SVM 時間 所以先註解掉 SVM 的部分
        temptime = time.process_time()
        meanDe = predictDe(train, test, allRandom)
        endDe = time.process_time()
        meanSVM = predictSVM(train, test, allRandom)
        endSVM = time.process_time()

        # write in excel prepare
        # 時間
        durDe = endDe - start
        durDe = round(durDe, 3)

        durSVM = endSVM - start - (endDe - temptime)
        durSVM = round(durSVM, 3)

        print('into the file')
        wb = load_workbook(
            '/Users/emily/Desktop/Research/fs_ensemble/of.xlsx')
        # sheet = wb['Ensemble 三個方法 Random']
        # sheet = wb['Ensemble 三個方法 ElbowCenter']
        # sheet Name
        sheet = wb[sheetName]
        print(fileName)
        rr = 0
        # rr = id +1
        sheet.cell(row=id+1, column=1, value=fileName)
        sheet.cell(row=id+1, column=cell, value=meanDe)
        # 時間
        # rr = id +2
        sheet.cell(row=2+id, column=cell, value=durDe)
        sheet.cell(row=3+id, column=cell, value=meanSVM)
        sheet.cell(row=4+id, column=cell, value=durSVM)
        # 找最大值 並上紅色
        '''
        ma = meanDe
        # ma = meanSVM
        tempappr.append(ma)

        fontRed = Font(color='FF0000', size=16)  # point at red font
        fontBoldRed = Font(color='FF0000', bold=True, size=16)
        if ma > maxsingle:

            sheet.cell(1+id, cell).font = fontBoldRed  # C4.5
            # sheet.cell(3+id, cell).font = fontBoldRed  # SVM

            print('meanDe233', ma)
        if index == 2 or index == 9 or index == 12:
            maxcell = max(tempappr)  # 每個方法的最大值
            if maxcell < maxsingle:
                print("new round3", tempappr)
                maxindex = tempappr.index(maxcell)  # 最大值的 index
            # print('maxindex', maxindex)
                # r = 3+id  # SVM
                r = 1 + id  # C4.5
                c = cell - (len(tempappr) - maxindex) + 1
                print('row3', r, '  column4', c)
                sheet.cell(r, c).font = fontRed

                print('meanDe111', maxcell)
            tempappr = []
        # 若大於 singel method 則標粗體
        '''
        wb.save('/Users/emily/Desktop/Research/fs_ensemble/of.xlsx')

    os.chdir(originpath)


# folder process
def train_test_split(folder):
    train = []
    test = []
    # os.chdir(folder)
    dirs = os.listdir(folder)
    for i in dirs:
        # print(i.split("-")[-1])
        if("xlsx" in i):
            if("tra" in i):
                train.append(i)

            elif("tst" in i):
                test.append(i)
    train = sorted(train)
    test = sorted(test)
    return train, test


print(os.getcwd())
# load data
df = pd.read_excel('./alldataName.xlsx')
df = pd.DataFrame(df)
dfa = df[df['#Attributes (R/I/N)'] >= 10]
print(dfa['Name '])

f1 = list(permutations("442", 3))
f2 = list(permutations("253", 3))
f3 = list(permutations("334", 3))

f = f1+f2+f3
temp = []
# get rid of the repeat ratio composition
for i in f:
    if i not in temp:
        temp.append(i)
f = temp
folderpath = []
id = 0
# for i in dfa['Name ']:

#     tmp = '../data/all/' + i
#     folderpath.append(tmp)

realname = []
for dir in os.listdir('../data/all'):
    realname.append(dir)
for i in dfa['Name ']:
    if i in realname:
        folderpath.append('../data/all/'+i)
    else:
        ll = '../data/all/' + i + '-5-fold'
        folderpath.append(ll)
print('fff', folderpath[3:14])
folderpath = folderpath[3:14]
folderpath = [
    '/Users/emily/Desktop/Research/oversampling_python/data/bank']
# folderpath = ['../data/all/vehicle1', '../data/all/vehicle3']

os.chdir('/Users/emily/Desktop/Research/oversampling_python/FS_ensemble')
for folder in folderpath:
    cell = 2

    path = folder

    train, test = train_test_split(folder)
    column = 2
    calculatethreemethod(train, test, id, path, 'elbowCenter',
                         'firstoversampling')

    print('===============')
    id = id+5
